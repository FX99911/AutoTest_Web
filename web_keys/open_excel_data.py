from openpyxl import load_workbook
from selenium.webdriver.common.by import By

from web_keys.montage_url import home




def read_excel(file_path):   #只需要输入相对路径就可以

    # 加载指定路径的Excel工作簿
    wb = load_workbook(f'{home}/{file_path}')
    # 获取工作簿中当前激活的工作表
    sheet = wb.active
    # 用于存储最终数据的字典，键为第一列的值（作为变量名），值根据情况可以是字符串或列表
    data_dict = {}
    # 遍历工作表中的每一行数据，min_row=1表示从第一行开始，values_only=True表示只获取单元格的值
    for row in sheet.iter_rows(min_row=1, values_only=True):
        # 获取当前行第一列的值，作为变量名
        var_name = row[0]
        # 如果变量名不为空
        if var_name:
            value_list = []
            # 从第三列（索引为2）开始遍历单元格
            for cell in row[2:]:
                # 如果单元格有值，则将其添加到value_list中
                if cell:
                    value_list.append(cell)
            # 判断值列表的长度
            if len(value_list) == 1:
                # 如果值列表只有一个元素，将该元素作为值（字符串格式）存入字典
                data_dict[var_name] = value_list[0]
            elif len(value_list) > 1:
                # 如果值列表有多个元素，将值列表存入字典
                data_dict[var_name] = value_list
            # 如果值列表为空，不进行处理，保持字典中该键对应的值为默认状态（这里是None，因为没赋值）
    # 返回存储数据的字典
    return data_dict


def convert_by_string(lst):
    """
    此函数用于将列表中以 'By.' 开头的字符串转换为 By 类对应的属性
    :param lst: 待处理的列表，列表元素可能包含以 'By.' 开头的字符串
    :return: 处理后的新列表，其中以 'By.' 开头的字符串已转换为 By 类的属性
    """
    # 初始化一个空列表，用于存储转换后的元素
    new_lst = []
    # 遍历输入列表中的每个元素
    for item in lst:
        # 检查元素是否为字符串类型，并且是否以 'By.' 开头
        if isinstance(item, str) and item.startswith('By.'):
            try:
                # 通过分割字符串，提取出定位方式的名称，例如 'id'
                locator_name = item.split('.')[-1]
                # 使用 getattr 函数从 By 类中获取对应的属性
                # 例如，如果 locator_name 为 'id'，则获取 By.ID
                new_lst.append(getattr(By, locator_name))
            except AttributeError:
                # 如果 By 类中不存在指定名称的属性，捕获 AttributeError 异常
                # 并打印错误信息，提示用户 By 类中没有该属性
                print(f"错误：By 类中不存在名为 {locator_name} 的属性。")
                # 将原字符串添加到新列表中，保持元素不变
                new_lst.append(item)
        else:
            # 如果元素不是以 'By.' 开头的字符串，直接添加到新列表中
            new_lst.append(item)
    # 返回处理后的新列表
    return new_lst


def convert_dict(dictionary):
    """
    此函数用于处理字典，将字典中值为列表的元素里以 'By.' 开头的字符串转换为 By 类对应的属性
    :param dictionary: 待处理的字典
    :return: 处理后的字典
    """
    new_dict = {}
    for key, value in dictionary.items():
        if isinstance(value, list):
            new_dict[key] = convert_by_string(value)
        else:
            new_dict[key] = value
    return new_dict



def excel_variables(file_path):
     # 使用原始字符串表示路径
     file_path = 'cases_date/test_excel.xlsx'
     # 调用read_excel函数读取Excel文件数据(字典)
     data_dict = read_excel(file_path)
     # 调用convert_dict函数把字典里的'By.xxx转为By属性'
     data_dict = convert_dict(data_dict)
     print(data_dict)
     return data_dict





file_path = 'cases_date/test_excel.xlsx'
excel_variables(file_path)