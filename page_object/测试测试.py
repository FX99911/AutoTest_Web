from openpyxl import load_workbook
from web_keys.montage_url import home
from selenium.webdriver.common.by import By


def read_excel(file_path):
    # 加载指定路径的Excel工作簿
    wb = load_workbook(f'{home}/{file_path}')
    # 获取工作簿中当前激活的工作表
    sheet = wb.active
    # 用于存储最终数据的字典，键为第一列的值（作为变量名），值根据情况可以是字符串或列表
    data_dict = {}
    # 遍历工作表中的每一行数据，min_row=1表示从第一行开始，values_only=True表示只获取单元格的值
    for row in sheet.iter_rows(min_row=1, values_only=True):
        # 获取当前行第一列的值，作为变量名
        var_name = row[0]
        # 如果变量名不为空
        if var_name:
            value_list = []
            # 从第三列（索引为2）开始遍历单元格
            for cell in row[2:]:
                # 如果单元格有值，则将其添加到value_list中
                if cell:
                    value_list.append(cell)
            # 判断值列表的长度
            if len(value_list) == 1:
                # 如果值列表只有一个元素，将该元素作为值（字符串格式）存入字典
                data_dict[var_name] = value_list[0]
            elif len(value_list) > 1:
                # 如果值列表有多个元素，将值列表存入字典
                data_dict[var_name] = value_list
            # 如果值列表为空，不进行处理，保持字典中该键对应的值为默认状态（这里是None，因为没赋值）
    # 返回存储数据的字典
    return data_dict


def str_to_by_locator(locator_str):
    """
    将字符串形式的定位方式转换为 By 类的属性
    :param locator_str: 表示定位方式的字符串，例如 'By.XPATH'
    :return: By 类的属性
    """
    try:
        # 提取定位方式名称
        locator_name = locator_str.split('.')[-1]
        # 获取 By 类的对应属性
        return getattr(By, locator_name)
    except AttributeError:
        print(f"错误：By 类中不存在名为 {locator_name} 的属性。")
        return None


def excel_variables():
    # 使用原始字符串表示路径
    file_path = 'cases_date/test_excel.xlsx'
    # 调用read_excel函数读取Excel文件数据
    result_dict = read_excel(file_path)
    print(result_dict)
    # 遍历结果字典，打印每个变量名及其对应的值（可能是字符串或列表）
    a = -1
    for key, value in result_dict.items():
        a += 1
        if isinstance(value, list):
            new_value = []
            for item in value:
                if isinstance(item, str) and item.startswith('By.'):
                    item = str_to_by_locator(item)
                new_value.append(item)
            result_dict[key] = new_value
        elif isinstance(value, str) and value.startswith('By.'):
            result_dict[key] = str_to_by_locator(value)
        print(f'{a}--{key}:{result_dict[key]}')

excel_variables()